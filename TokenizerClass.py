from nltk.stem.snowball import SnowballStemmer
from nltk.tokenize import word_tokenize
from nltk.util import ngrams
from nltk.corpus import stopwords
from nltk.stem.porter import *
from openpyxl import Workbook
import mysql.connector
import os
import sys
import math
import time



def getDirectoryOfData():
    '''
        This function will get the 'data' directory of current folder
    '''
    dataDir = os.path.realpath('.') +'/data'
    dataFiles = os.listdir(dataDir)
    filesPath = []
    for files in dataFiles:
        filesPath.append(dataDir+"/"+files)
    dataDir = os.path.realpath('.') + '/sentence'
    dataFiles = os.listdir(dataDir)
    for files in dataFiles:
        filesPath.append(dataDir+"/"+files)
    # print(filesPath)
    return filesPath

class Tokenizer(object):
    '''
        This class will store all the data into a spreadsheet
        instead of a console or database.
    '''

    def __init__(self,file_paths,excel_sheet):
        '''
            This initializes the Tokenizer Class

            Class Variables:
            __tokenized_words {} and __key_words {} format:
                { 'Token1': [ { DocumentID : [ [TokenPos,TokenPos,...], TF, TFIDF ] },  DF] }
            __variation_of_stem_forms {} format:
                { 'Stem Word' : [Variation1, Variation2, Variation3...] }


            Parameters:
                file_paths (String): Array of document path
                excel_sheet (Workbook): Spreadsheet to input data
        '''
        self.__tokenized_words = {}
        self.__key_words = {}
        self.__tokenized_stemmed_words = {}
        self.__variation_of_stem_forms = {}
        self.__tfidf = []
        self.data_excel_sheet = excel_sheet
        self.__file_paths = file_paths
        self.__total_num_of_doc = len(file_paths)

        self.__dbUser = "root"
        self.__dbPassword = ""
        self.__dbDatabase = "db_stem"
        self.max_gap = 0
        self.__dbHost = "localhost"







    def run(self):
        '''
            This runs the tokenizer
        '''
        # This will collect all of the TF first
        for docID,path in enumerate(self.__file_paths):
            tokenized_stemmed_words = self.start_stemming_document(path)
            # self.store_data_into_dict(tokenized_text, docID, self.__tokenized_words)
            # self.store_data_into_dict(tokenized_stemmed_words, docID, self.__tokenized_words)
            self.store_data_into_dict(tokenized_stemmed_words, docID, self.__tokenized_stemmed_words)

        # This will collect DF and TFiDF
        for token_value in self.__tokenized_stemmed_words.values():
            num_of_doc_with_token = len(token_value[0])
            self.store_df_calc(token_value,self.__total_num_of_doc,num_of_doc_with_token)
            self.store_tfidf_calc(token_value)
        # self.__store_data_into_excel(self.__tokenized_words)
        # self.__store_data_into_db(self.__tokenized_stemmed_words)
        # self.__store_data_into_excel(self.__tokenized_stemmed_words)
        # self.__calculate_max_gap()
        self.__calculate_max_gap_wo_db()
        self.max_gap = self.max_gap/2
        for token,values in self.__tokenized_stemmed_words.items():
            for index,(key,value) in enumerate(values[0].items()):
                if(value[2] >= self.max_gap and value[2] <= 0.09):
                    self.__key_words[str(index)] = token
        for key,value in self.__tokenized_stemmed_words.items():
            print("Token: [" + str(key) + "] ||| Value: [" + str(value)+"]")


    def getTFIDFList(self):
        return self.__tfidf


    def __calculate_max_gap_wo_db(self):
        '''
            Calculating the max gap without the database so it's faster
        '''
        for token,values in self.__tokenized_stemmed_words.items():
            for key,value in values[0].items():
                self.__tfidf.append(value[2])

        previous = 0
        print(self.__tfidf.sort())
        for value in self.__tfidf:
            if previous == 0:
                previous = value
            else:
                gap = abs(value - previous)
                if(self.max_gap < gap):
                    self.max_gap = gap
                previous = value




    def start_stemming_document(self,file_path_of_doc):
        '''
            Stems the text for each word from the document

            Parameters:
                file_path (String): the current document path to perform stem operation

            Return:
                (List): an array of tokenized words
        '''
        ps = PorterStemmer()
        with open(file_path_of_doc,"r",encoding="UTF-8",errors='ignore') as document:
            text = document.read()
        tokenized_text = word_tokenize(text)
        stemmed_tokenized_text = [ps.stem(words) for words in tokenized_text]
        bigrm = ngrams(tokenized_text,2)
        trigrm = ngrams(tokenized_text,3)
        fourgrm = ngrams(tokenized_text,4)
        fivegrm = ngrams(tokenized_text,5)
        fused_grm = list(bigrm)+list(trigrm)+list(fourgrm)+list(fivegrm)

        grm_list=[]
        for tuple in fused_grm:
            string = ' '.join(tuple)
            grm_list.append(string)
        wordlist = stemmed_tokenized_text + grm_list
        return wordlist






    def start_tokenizing_document(self,file_path_of_doc):
        '''
            Tokenizes the text for each word from the document

            Parameters:
                file_path (String): the current document path to perform stem operation

            Return:
                (List): an array of tokenized words
        '''
        with open(file_path_of_doc,"r") as document:
            text = document.read()
        return word_tokenize(text)






    def store_data_into_dict(self, stemmed_text, doc_ID, word_dict = {}):
        '''
            Store stem word, Token position, document id, and TF into class variable dictionary
            What is stored during this operation:
                {  'Token' : [{ DocumentID : [[TokenPos,TokenPos,...], TF] }   ]   }

            Parameters:
                file_path (String): the current document path to perform stem operation
                stemmed_text (List): the array of tokenized words
                doc_ID (int): the current document ID
            Return:
                (Dictionary): Return a dictionary of values in the format above
        '''
        current_doc_word_count = len(stemmed_text)
        for tokenPos, word in enumerate(stemmed_text):
            if word in word_dict:
                value = word_dict[word][0]
                if doc_ID in value:
                    value[doc_ID][0].append(tokenPos)
                else:
                    value[doc_ID] = [[tokenPos]]
            else:
                word_dict[word] = [{ doc_ID : [[tokenPos]] }]
        for doc in word_dict.values():
            length_of_doc = len(doc[0])
            for currentDoc in doc[0].values():
                if len(currentDoc) != 2:
                    self.store_tf_calc(currentDoc,current_doc_word_count)
                else:
                    continue
        return word_dict







    def store_tf_calc(self, document_id_value,doc_word_count):
        '''
            Stores the token frequency (TF) in given dictionary
            Before:
                {  'Token' : [{ DocumentID : [[TokenPos,TokenPos,...]]  }       }]

            After:
                {  'Token' : [{ DocumentID : [[TokenPos,TokenPos,...], TF] }     }]

            Parameters:
                document_id_value (Dictionary): the DocumentID value
        '''
        total_tokens = len(document_id_value[0])
        document_id_value.append(self.__calculate_tf(total_tokens,doc_word_count))








    def store_tfidf_calc(self, document_id_value):
        '''
            Stores the token frequency (TF) in given dictionary
            Before:
                {  'Token' : [ { DocumentID : [[TokenPos,TokenPos,...], TF] }, DF  ]   }

            After:
                {  'Token' : [   {DocumentID : [[TokenPos,TokenPos,...], TF,TFIDF]}     ,      DF     ]   }

            Parameters:
                document_id_value (Dictionary): the DocumentID value
        '''
        for current_doc_values in document_id_value[0].values():
            tf = current_doc_values[1]
            df = document_id_value[1]
            current_doc_values.append(self.__calculate_tfidf(tf,df))







    def store_df_calc(self,document_id_value,total_num_of_doc, num_of_doc_with_token):
        '''
            Stores the document frequency (DF) in given dictionary per token

            Before:
                {  'Token' : [ { DocumentID : [[TokenPos,TokenPos,...], TF] }   ]   }

            After:
                {  'Token' : [ { DocumentID : [[TokenPos,TokenPos,...], TF] }, DF ] }

            Parameters:
                document_id_value (List): the DocumentID value containing the List for token
        '''
        document_id_value.append(self.__calculate_df(total_num_of_doc,num_of_doc_with_token))









    def __calculate_tf(self,total_tokens,document_word_count):
        '''
            Calculates the token frequency (TF) in a document

            Parameters:
                total_tokens (int): the total amount of that specific word
                in a document that appeared

                document_word_count (int): the total word count in current document

            Return:
                (int) : Token Frequency Calculation
        '''
        return total_tokens/document_word_count





    def __calculate_df(self,total_num_of_doc, num_of_doc_with_token):
        '''
            Calculates the document frequency (DF)

            Parameters:
                total_num_of_doc (int): the total amount of documents

                num_of_doc_with_tokens (int): the total documents containing the token

            Return:
                (int) : Document frequency calculation
        '''
        return math.log(total_num_of_doc/num_of_doc_with_token) * 1.0





    def __calculate_tfidf(self,tf_value, df_value):
        '''
            Calculates the TFiDF to find key words

            Parameters:
                tf_value (int): Token frequency
                df_value (int): Document Frequency

            Return:
                (int) : TFiDF Calculation
        '''
        return tf_value * df_value





    def __store_data_into_excel(self,word_dict={}):
        '''
            Stores value into Excel Spreadsheet

            Parameters:
                word_dict (Dictionary): the data of all the correctly formatted as shown in init
        '''
        count_row = 1
        self.data_excel_sheet.cell(row = count_row,column = 1, value="token")
        self.data_excel_sheet.cell(row = count_row,column = 2, value="document_ID")
        self.data_excel_sheet.cell(row = count_row,column = 3, value="token_Position")
        self.data_excel_sheet.cell(row = count_row,column = 4, value="tf")
        self.data_excel_sheet.cell(row = count_row,column = 5, value="df")
        self.data_excel_sheet.cell(row = count_row,column = 6, value="tfidf")
        count_row = count_row + 1

        sorted_word_dict = {}
        for k in sorted(word_dict):
            test = "%s" % k
            if test.replace("'", "").isalpha():
                sorted_word_dict[k] = word_dict[k]

        for token,documents in sorted_word_dict.items():
            for key,val in documents[0].items():
                # print("Token: " +str(token) + " Value: "+ str(val))
                document_ID = key
                token_Position = val[0]
                tf = val[1]
                df = documents[1]
                tfidf = val[2]
                self.data_excel_sheet.cell(row = count_row,column = 1, value=token)
                self.data_excel_sheet.cell(row = count_row,column = 2, value=document_ID)
                self.data_excel_sheet.cell(row = count_row,column = 3, value=str(token_Position))
                self.data_excel_sheet.cell(row = count_row,column = 4, value=tf)
                self.data_excel_sheet.cell(row = count_row,column = 5, value=df)
                self.data_excel_sheet.cell(row = count_row,column = 6, value=tfidf)
                count_row = count_row + 1







    def __store_data_into_db(self,word_dict={}):
        '''
            Stores value into database

            Parameters:
                word_dict (Dictionary): the data of all the correctly formatted as shown in init
        '''
        mydb = mysql.connector.connect(
                                     user = self.__dbUser,
                                     password = self.__dbPassword,
                                     host = self.__dbHost)
        mycursor = mydb.cursor()
        self.recreateTable(mycursor,mydb)
        mycursor.close()
        mydb.close()

        mydb = mysql.connector.connect(
                                     user = self.__dbUser,
                                     password = self.__dbPassword,
                                     host = self.__dbHost,
                                     database = self.__dbDatabase)
        mycursor = mydb.cursor(buffered=True)

        past_millis = int(round(time.time() * 1000))

        for token,documents in word_dict.items():
            for key,val in documents[0].items():
                document_ID = key
                token_Position = val[0]
                tf = val[1]
                df = documents[1]
                tfidf = val[2]
                self.updateDatabase(mycursor,mydb, token, document_ID,tf,df,tfidf)
        #
        # mycursor.execute("SELECT token, tfidf FROM stem_data ORDER BY tfidf DESC")
        # row = mycursor.fetchone()
        # previous = 0
        # while row is not None:
        #     if previous == 0:
        #         print("Token: " + row[0] + ", Gap: 0")
        #     else:
        #         print("Token: " + row[0] + ", Gap: " + str((previous - row[1])))
        #     previous = row[1]
        #     row = mycursor.fetchone()

        mycursor.close()
        mydb.close()
        current_millis = int(round(time.time() * 1000))
        result = current_millis-past_millis
        print("Milliseconds: "+ str(result))



    def recreateTable(self,cursor,mydb):
        '''
            Recreates database and table if it doesn't exist

            Parameters:
                cursor (DB cursor Object): The database cursor to query and execute commands
                mydb (MySQL Object): the database connection object
        '''
        try:
            cursor.execute("DROP DATABASE db_stem")
            mydb.commit()
        except:
            pass
        try:
            cursor.execute("CREATE DATABASE IF NOT EXISTS db_stem CHARACTER SET utf8 COLLATE utf8_unicode_ci")
            mydb.commit()
        except:
            pass
        cursor.close()
        mydb.close()
        try:
            mydb = mysql.connector.connect(
                                        user = self.__dbUser,
                                        password = self.__dbPassword,
                                        host = self.__dbHost,
                                        database=self.__dbDatabase)
            cursor = mydb.cursor()
        except:
            print("Connection failed to login")
        try:
            cursor.execute("DROP TABLE np_stem_data_t1")
            cursor.execute("DROP TABLE np_stem_data_t2")
            mydb.commit()
        except:
            print("Query has failed to drop the table")
        try:
            cursor.execute("CREATE TABLE np_stem_data_t1(token VARCHAR(60), df DECIMAL(11,10) NOT NULL, PRIMARY KEY (token)) ENGINE=InnoDB CHARACTER SET utf8 COLLATE utf8_unicode_ci")
            cursor.execute("CREATE TABLE np_stem_data_t2(token VARCHAR(60), doc_ID INT NOT NULL,tf DECIMAL(11,10) NOT NULL, tfidf DECIMAL(11,10) NOT NULL, FOREIGN KEY (token) REFERENCES np_stem_data_t1(token)) ENGINE=InnoDB CHARACTER SET utf8 COLLATE utf8_unicode_ci")
            mydb.commit()
        except:
            print("Query has failed to create the table")




    def updateDatabase(self,cursor,mydb, token, document_ID,tf,df,tfidf):
        '''
            Insert values into tables

            Parameters:
                cursor (DB cursor Object): The database cursor to query and execute commands
                mydb (MySQL Object): the database connection object
                token, document_ID, tf, df, tfidf: The values from the dictionary
        '''
        esc_token = token.replace("'","''")
        # print(token)
        check_data = "SELECT (1) FROM np_stem_data_t1 WHERE token ='%s' limit 1" % (esc_token)
        cursor.execute(check_data)
        results = cursor.fetchone()
        if not results:
            sql="INSERT INTO np_stem_data_t1(token, df) VALUES (%s,%s)"
            value = (token,df)
            cursor.execute(sql,value)
        sql="INSERT INTO np_stem_data_t2(token, doc_ID,tf, tfidf) VALUES (%s,%s,%s,%s)"
        value = (token,document_ID,tf,tfidf)
        cursor.execute(sql,value)
        check_data = "SELECT token,df FROM np_stem_data_t1 WHERE token = '%s'" % (esc_token)
        mydb.commit()



    def print_data(self):
        print(self.__tokenized_words)





    def get_tokenized_words_data(self):
        return self.__tokenized_words





    def get_key_words_data(self):
        return self.__key_words



    def __calculate_max_gap(self):
        '''
            This will get the max gap for all the stemmed words in the documents
        '''
        try:
            mydb = mysql.connector.connect(
                                     user = self.__dbUser,
                                     password = self.__dbPassword,
                                     host = 'localhost',
                                     database = self.__dbDatabase)
            mycursor = mydb.cursor(buffered=True)
        except:
            print("Database connection has failed to connect")
        try:
            mycursor.execute("SELECT tfidf FROM np_stem_data_t2 ORDER BY tfidf ASC")
            results = mycursor.fetchall()
            previous = 0
            for row in results:
                if previous == 0:
                    previous = row[0]
                else:
                    gap = abs(row[0] - previous)
                    print("Gap: %.9f" % gap)
                    if(self.max_gap < gap):
                        self.max_gap = gap
                        print("New max gap: %.9f" % gap)
                    previous = row[0]

        except:
            print("Query tfidf has failed to calculate max gap")
        mycursor.close()
        mydb.close()





    def get_max_gap(self):
        '''
            This will get the max gap from the tfidf values, calculated
            from __calculate_max_gap function

            Return:
                max_gap (Float): This will return the max gap value
        '''
        return self.max_gap

    def run_gap(self):
        self.__calculate_max_gap()





def main():
    '''
        Runs the entire tokenizer process.
    '''
    data_Files_Path = getDirectoryOfData()
    wb = Workbook()
    data_excel_sheet = wb.active


    tokenizer1 = Tokenizer(data_Files_Path,data_excel_sheet)
    tokenizer1.run()
    result = tokenizer1.get_max_gap()
    print("The Max Gap is : " + '{0:.11g}'.format(result))
    print(sorted(tokenizer1.get_key_words_data().values()))
    # print(tokenizer1.getTFIDFList())
    # wb.save('Tokenizer_data.xlsx')

main()
